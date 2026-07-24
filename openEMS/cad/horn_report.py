#!/usr/bin/env python3
"""
Horn antenna report builder.

Reads the CSVs written by horn_run.m (gain-vs-frequency, S11-vs-frequency,
E/H-plane radiation pattern at f0) and horn_wallloss.py (coating comparison),
computes the standard antenna figures of merit, and writes:

  horn_results/horn_metrics.csv      key,value  (HPBW, SLL, F/B, effs, ...)
  horn_results/horn_results.xlsx     paper-style multi-sheet data file

The metric set mirrors what "MXene-Coated 3D Printed Horn Antennas for Ku
Frequency Band" (Sarpanah Sourkouhi et al., Adv. Mater. Technol. 2026,
DOI 10.1002/admt.202502239) reports: S11 vs frequency, gain vs frequency,
gain summary (average/center/minimum), and E-/H-plane radiation patterns.

Run (after horn_run.m + horn_wallloss.py):
    /opt/anaconda3/bin/python horn_report.py [tag]        # tag defaults to 'pec'
Test with mock data:
    HORN_RESULTS_DIR=/tmp/mock /opt/anaconda3/bin/python horn_report.py pec
"""
import os, sys, csv, math

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.environ.get("HORN_RESULTS_DIR") or os.path.join(HERE, "horn_results")
TAG  = (sys.argv[1] if len(sys.argv) > 1 else "pec")


# ----------------------------- small CSV helpers -----------------------------
def read_rows(path):
    """Return (header:list, rows:list[list[float|str]]) or (None, [])."""
    if not os.path.isfile(path):
        return None, []
    with open(path, newline="") as f:
        rd = list(csv.reader(f))
    if not rd:
        return None, []
    header, body = rd[0], rd[1:]
    out = []
    for r in body:
        if not any(c.strip() for c in r):
            continue
        conv = []
        for c in r:
            try:
                conv.append(float(c))
            except ValueError:
                conv.append(c)
        out.append(conv)
    return header, out


def col(header, rows, name):
    if header is None or name not in header:
        return []
    i = header.index(name)
    return [r[i] for r in rows]


# --------------------------- pattern figures of merit ------------------------
def hpbw(theta, gain_dbi):
    """Half-power (-3 dB) beamwidth in degrees, around the boresight peak."""
    n = len(gain_dbi)
    ip = max(range(n), key=lambda i: gain_dbi[i])
    half = gain_dbi[ip] - 3.0

    def crossing(i_out, i_in):
        g0, g1 = gain_dbi[i_out], gain_dbi[i_in]
        t0, t1 = theta[i_out], theta[i_in]
        if g1 == g0:
            return t1
        return t0 + (half - g0) * (t1 - t0) / (g1 - g0)

    iL = ip
    while iL > 0 and gain_dbi[iL - 1] >= half:
        iL -= 1
    thL = theta[iL] if iL == 0 else crossing(iL - 1, iL)
    iR = ip
    while iR < n - 1 and gain_dbi[iR + 1] >= half:
        iR += 1
    thR = theta[iR] if iR == n - 1 else crossing(iR + 1, iR)
    return abs(thR - thL)


def sidelobe_level(theta, gain_dbi):
    """First/worst sidelobe level in dB below the main-beam peak (positive)."""
    n = len(gain_dbi)
    ip = max(range(n), key=lambda i: gain_dbi[i])
    pk = gain_dbi[ip]

    def side(indices):
        # walk away from the peak, past the first null (local minimum),
        # then take the highest lobe beyond it
        i = 1
        while i < len(indices) - 1 and gain_dbi[indices[i]] <= gain_dbi[indices[i - 1]]:
            i += 1
        if i >= len(indices) - 1:
            return None
        return pk - max(gain_dbi[j] for j in indices[i:])

    cands = [x for x in (side(list(range(ip, n))), side(list(range(ip, -1, -1)))) if x is not None]
    return min(cands) if cands else None    # smallest drop = strongest sidelobe


def front_to_back(theta, gain_dbi):
    """Peak gain minus gain 180 deg away (dB)."""
    n = len(gain_dbi)
    ip = max(range(n), key=lambda i: gain_dbi[i])
    back = theta[ip] + 180.0
    if back > 180.0:
        back -= 360.0
    ib = min(range(n), key=lambda i: abs(theta[i] - back))
    return gain_dbi[ip] - gain_dbi[ib]


# --------------------------------- load data ---------------------------------
def geo():
    try:
        vals = [float(v) for v in open(os.path.join(OUT, "horn_geo.txt")).read().split()]
        keys = ["a", "b", "Ax", "By", "feed", "flare", "f0"]
        return dict(zip(keys, vals))
    except Exception:
        return {}


G = geo()
f0 = G.get("f0", None)

gh, grows = read_rows(os.path.join(OUT, f"horn_gain_{TAG}.csv"))
sh, srows = read_rows(os.path.join(OUT, f"horn_s11_{TAG}.csv"))
ph, prows = read_rows(os.path.join(OUT, f"horn_pattern_{TAG}.csv"))
ch, crows = read_rows(os.path.join(OUT, "horn_coating_loss.csv"))

metrics = {}

# gain-vs-frequency summary (matches paper Fig 1e: average / center / minimum)
gfreq = col(gh, grows, "freq_ghz")
gd = col(gh, grows, "Dmax_dBi")
gs11 = col(gh, grows, "s11_db")
if gd:
    metrics["directivity_avg_dBi"] = sum(gd) / len(gd)
    metrics["directivity_min_dBi"] = min(gd)
    metrics["directivity_max_dBi"] = max(gd)
    if f0 is not None and gfreq:
        kc = min(range(len(gfreq)), key=lambda i: abs(gfreq[i] - f0 / 1e9))
        metrics["directivity_center_dBi"] = gd[kc]
        for name in ("rad_eff_pct", "aperture_eff_pct", "s11_db"):
            c = col(gh, grows, name)
            if c:
                metrics[name.replace("_pct", "_pct_f0").replace("_db", "_db_f0")] = c[kc]

# pattern figures of merit at f0
theta = col(ph, prows, "theta_deg")
gE = col(ph, prows, "gain_Eplane_dBi")
gH = col(ph, prows, "gain_Hplane_dBi")
if theta and gE:
    metrics["peak_gain_dBi"] = max(max(gE), max(gH) if gH else -999)
    metrics["HPBW_E_deg"] = hpbw(theta, gE)
    metrics["SLL_E_dB"] = sidelobe_level(theta, gE)
    metrics["FB_E_dB"] = front_to_back(theta, gE)
    if gH:
        metrics["HPBW_H_deg"] = hpbw(theta, gH)
        metrics["SLL_H_dB"] = sidelobe_level(theta, gH)
        metrics["FB_H_dB"] = front_to_back(theta, gH)

# S11-derived: -10 dB impedance bandwidth, VSWR @ f0, input impedance @ f0,
# mismatch efficiency and realized gain (directivity incl. reflection loss)
sfreq = col(sh, srows, "freq_ghz")
sdb = col(sh, srows, "s11_db")
if sfreq and sdb:
    below = [sfreq[i] for i in range(len(sfreq)) if sdb[i] < -10.0]
    if below:
        bw = max(below) - min(below)
        metrics["bandwidth_S11_10dB_GHz"] = bw
        if f0:
            metrics["fractional_bw_pct"] = bw / (f0 / 1e9) * 100.0
    if f0:
        kc = min(range(len(sfreq)), key=lambda i: abs(sfreq[i] - f0 / 1e9))
        mag = 10 ** (sdb[kc] / 20.0)
        metrics["VSWR_f0"] = (1 + mag) / max(1 - mag, 1e-6)
        metrics["mismatch_eff_pct"] = (1 - mag ** 2) * 100.0
        zr = col(sh, srows, "Zin_real_ohm"); zi = col(sh, srows, "Zin_imag_ohm")
        if zr and kc < len(zr): metrics["Zin_real_ohm_f0"] = zr[kc]
        if zi and kc < len(zi): metrics["Zin_imag_ohm_f0"] = zi[kc]
        if "directivity_center_dBi" in metrics:
            metrics["realized_gain_total_dBi"] = metrics["directivity_center_dBi"] + 10 * math.log10(max(1 - mag ** 2, 1e-6))

# write horn_metrics.csv
os.makedirs(OUT, exist_ok=True)
with open(os.path.join(OUT, "horn_metrics.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["metric", "value"])
    for k, v in metrics.items():
        w.writerow([k, ("" if v is None else round(v, 4))])
print("wrote", os.path.join(OUT, "horn_metrics.csv"))
for k, v in metrics.items():
    print(f"  {k:<24} {'' if v is None else round(v,3)}")


# --------------------------------- Excel -------------------------------------
def build_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception as e:
        print(f"(openpyxl not available, skipping .xlsx: {e})")
        return None
    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0B6E54")
    title_font = Font(bold=True, size=12, color="0B6E54")

    def sheet_from_csv(title, header, rows):
        ws = wb.create_sheet(title)
        if header:
            ws.append(header)
            for c in ws[1]:
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for col_cells in ws.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 32)
        return ws

    # Summary sheet
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Horn antenna — openEMS full-wave results"; ws["A1"].font = title_font
    ws.append([])
    ws.append(["Quantity", "Value", "Unit"])
    for c in ws[3]:
        c.font = hdr_font; c.fill = hdr_fill
    def line(q, v, u=""):
        ws.append([q, ("" if v is None else (round(v, 4) if isinstance(v, float) else v)), u])
    if G:
        line("Feed a × b", f"{G.get('a','?')} × {G.get('b','?')}", "mm")
        line("Aperture Ax × By", f"{G.get('Ax','?')} × {G.get('By','?')}", "mm")
        line("Flare length", G.get("flare"), "mm")
        line("Center frequency f0", (f0 / 1e9 if f0 else None), "GHz")
    order = [("directivity_center_dBi", "Gain @ f0 (center)", "dBi"),
             ("directivity_avg_dBi", "Gain (band average)", "dBi"),
             ("directivity_max_dBi", "Gain (band maximum)", "dBi"),
             ("directivity_min_dBi", "Gain (band minimum)", "dBi"),
             ("HPBW_E_deg", "HPBW E-plane", "deg"),
             ("HPBW_H_deg", "HPBW H-plane", "deg"),
             ("SLL_E_dB", "First sidelobe (E-plane)", "dB below peak"),
             ("SLL_H_dB", "First sidelobe (H-plane)", "dB below peak"),
             ("FB_E_dB", "Front-to-back (E-plane)", "dB"),
             ("FB_H_dB", "Front-to-back (H-plane)", "dB"),
             ("aperture_eff_pct_f0", "Aperture efficiency @ f0", "%"),
             ("rad_eff_pct_f0", "Radiation efficiency @ f0 (PEC)", "%"),
             ("s11_db_f0", "S11 @ f0", "dB"),
             ("VSWR_f0", "VSWR @ f0", ""),
             ("bandwidth_S11_10dB_GHz", "-10 dB impedance bandwidth", "GHz"),
             ("fractional_bw_pct", "Fractional bandwidth", "%"),
             ("Zin_real_ohm_f0", "Input impedance Re @ f0", "ohm"),
             ("Zin_imag_ohm_f0", "Input impedance Im @ f0", "ohm"),
             ("mismatch_eff_pct", "Mismatch efficiency @ f0", "%"),
             ("realized_gain_total_dBi", "Realized gain @ f0 (incl. mismatch)", "dBi")]
    for key, label, unit in order:
        if key in metrics:
            line(label, metrics[key], unit)
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=12)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 12), 40)

    if sh: sheet_from_csv("S11_vs_freq", sh, srows)
    if gh: sheet_from_csv("Gain_vs_freq", gh, grows)
    if ph: sheet_from_csv("Pattern_f0", ph, prows)
    if ch: sheet_from_csv("Coating_comparison", ch, crows)

    # one sheet PER COATING (all from the single validated PEC solve).
    # realized gain(f) = directivity(f) + that coating's gain penalty (dB).
    if ch and crows:
        idx = {name: i for i, name in enumerate(ch)}
        def cv(row, name):
            i = idx.get(name)
            return row[i] if (i is not None and i < len(row)) else None
        for row in crows:
            mat = str(cv(row, "material") or "coating")
            safe = "".join(c for c in mat if (c.isalnum() or c in " _-")).strip()[:24] or "coating"
            ws = wb.create_sheet(("Coat_" + safe)[:31])
            pen = cv(row, "gain_penalty_dB") or 0.0
            hdr = [("Coating", mat),
                   ("Conductivity (S/m)", cv(row, "sigma_S_per_m")),
                   ("Surface resistance Rs (ohm)", cv(row, "Rs_ohm")),
                   ("Radiation efficiency (%)", cv(row, "rad_eff_pct")),
                   ("Gain penalty vs PEC (dB)", cv(row, "gain_penalty_dB")),
                   ("Gain @ f0 (dBi)", cv(row, "gain_dBi"))]
            for q, v in hdr:
                ws.append([q, v])
            ws.append([])
            ws.append(["freq_ghz", "directivity_dBi", "realized_gain_dBi", "s11_db"])
            for c in ws[len(hdr) + 2]:
                c.font = hdr_font; c.fill = hdr_fill
            for i, f in enumerate(gfreq):
                d = gd[i] if i < len(gd) else None
                s = gs11[i] if i < len(gs11) else None
                ws.append([f, d, (round(d + pen, 3) if d is not None else None), s])
            for cc in ws.columns:
                wdt = max((len(str(c.value)) for c in cc if c.value is not None), default=12)
                ws.column_dimensions[cc[0].column_letter].width = min(max(wdt + 2, 12), 34)

    # Readme / provenance sheet
    rm = wb.create_sheet("Readme")
    notes = [
        ["Sheet", "Contents"],
        ["Summary", "Key figures of merit for this run (dimensions, gain, HPBW, sidelobe, F/B, efficiencies, S11)."],
        ["S11_vs_freq", "Return loss |S11| in dB vs frequency (port reflection)."],
        ["Gain_vs_freq", "Max directivity (dBi), radiated power, radiation & aperture efficiency, S11, per frequency."],
        ["Pattern_f0", "E-plane (phi=0) and H-plane (phi=90) gain patterns in dBi vs theta, at f0."],
        ["Coating_comparison", "Radiation efficiency / gain penalty per inner-wall coating (surface-impedance method)."],
        ["", ""],
        ["Method", "openEMS FDTD, PEC walls + near-field-to-far-field transform for gain/pattern."],
        ["Wall loss", "Surface-impedance perturbation from the PEC field: P_c=(Rs/2)∮|H_tan|^2 dA, Rs=sqrt(pi f mu0/sigma)."],
        ["Caveats", "openEMS S21/loss magnitude is coarse-mesh approximate; use for pattern/reflection. Coating changes gain <0.05 dB."],
        ["Metric set", "Mirrors Sarpanah Sourkouhi et al., Adv. Mater. Technol. 2026, DOI 10.1002/admt.202502239."],
    ]
    for r in notes:
        rm.append(r)
    for c in rm[1]:
        c.font = hdr_font; c.fill = hdr_fill
    rm.column_dimensions["A"].width = 22
    rm.column_dimensions["B"].width = 95

    path = os.path.join(OUT, "horn_results.xlsx")
    wb.save(path)
    print("wrote", path)
    return path


build_excel()
